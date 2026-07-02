# coding=utf-8

# --- Import required libraries ---
import os
import re
import time
import warnings
import sys
import cv2
import pytesseract
from PIL import Image
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# --- Set Tesseract OCR path ---
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# --- Ignore warnings ---
if not sys.warnoptions:
    warnings.simplefilter("ignore")
url = ""    

# --- Get user input from CLI or arguments ---
def get_inputs():

    if len(sys.argv) == 8:

        college = sys.argv[1].upper()
        year = sys.argv[2]
        branch = sys.argv[3].upper()
        low = int(sys.argv[4])
        high = int(sys.argv[5]) + 1
        semc = sys.argv[6]
        url = sys.argv[7]

    else:

        college = input("Enter the college code: ").upper()
        year = input("Enter the year: ")
        branch = input("Enter the branch: ").upper()
        low = int(input("Enter starting USN: "))
        high = int(input("Enter last USN: ")) + 1
        semc = input("Enter the Semester: ")
        url = input("Enter VTU Result URL: ")

    return college, year, branch, low, high, semc, url

# --- Set result URL based on semester ---
# if semc == "1":
#    url = "https://results.vtu.ac.in/DJcbcs24/index.php"
# elif semc == "4":
#    url = "https://results.vtu.ac.in/JJEcbcs25/index.php"
# else:
#    print(f"⚠️ Semester {semc} not supported yet. Exiting...")
#    sys.exit()

# --- Setup Selenium Chrome driver ---
service = Service(ChromeDriverManager().install())
options = webdriver.ChromeOptions()
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--start-maximized")
options.add_argument("--headless")  # Run in headless mode (no window)
driver = webdriver.Chrome(service=service, options=options)

# --- Create folder for captcha images ---
os.makedirs("captcha_dataset", exist_ok=True)

# --- Clean OCR output for CAPTCHA ---
def clean_captcha(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", text.strip())

# --- Solve CAPTCHA using OCR ---
def solve_captcha(image_path: str) -> str:
    img_cv = cv2.imread(image_path)
    gray = cv2.cvtColor(img_cv, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (3, 3), 0)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)
    cv2.imwrite("cap_processed.png", thresh)
    img = Image.open("cap_processed.png")
    captcha_text = pytesseract.image_to_string(
        img,
        config="--oem 3 --psm 7 -c tessedit_char_whitelist=abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789",
    )
    return clean_captcha(captcha_text)

# --- Fetch result for a single USN ---
# --- Fetch result for a single USN ---
def fetch_result(usn: str):

    while True:

        try:

            driver.get(url)

            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.NAME, "lns"))
            )

            # -------------------------
            # Enter USN
            # -------------------------

            usn_input = driver.find_element(By.NAME, "lns")
            usn_input.clear()
            usn_input.send_keys(usn)

            # -------------------------
            # Solve Captcha
            # -------------------------

            captcha_img = driver.find_element(
                By.XPATH,
                "//img[@alt='CAPTCHA code']"
            )

            captcha_img.screenshot("cap.png")

            captcha_text = solve_captcha("cap.png")

            if len(captcha_text) < 4:
                print("OCR Failed...")
                continue

            print(f"OCR detected captcha for {usn}: {captcha_text}")

            cap_input = driver.find_element(By.NAME, "captchacode")
            cap_input.clear()
            cap_input.send_keys(captcha_text)

            driver.find_element(By.ID, "submit").click()

            time.sleep(2)

            # -------------------------
            # Handle Alerts
            # -------------------------

            try:

                alert = driver.switch_to.alert

                txt = alert.text

                alert.accept()

                if "Invalid captcha" in txt:

                    print("❌ Invalid captcha...Retrying")
                    continue

                if "University Seat Number" in txt:

                    print(f"No Result : {usn}")
                    return None

            except:
                pass

            # -------------------------
            # Parse HTML
            # -------------------------

            soup = BeautifulSoup(driver.page_source, "html.parser")

            with open("page.html", "w", encoding="utf-8") as f:
                f.write(driver.page_source)

            student = {}

            # -------------------------
            # Student Details
            # -------------------------

            info_rows = soup.select(".divTableRow")

            for row in info_rows:

                cells = row.select(".divTableCell")

                if len(cells) != 2:
                    continue

                key = cells[0].get_text(" ", strip=True)
                value = cells[1].get_text(" ", strip=True)

                if "University Seat Number" in key:
                    student["USN"] = value

                elif "Student Name" in key:
                    student["Name"] = value

            if "USN" not in student:

                print(f"⚠️ Couldn't read {usn}. Retrying...")

                time.sleep(2)

                continue

            print(student)
                        # -------------------------
            # Semester
            # -------------------------

            semester = ""

            sem_tag = soup.find("b", string=lambda x: x and "Semester" in x)

            if sem_tag:
                semester = sem_tag.get_text(strip=True)

            student["Semester"] = semester

            # -------------------------
            # Subject Details
            # -------------------------

            subjects = []

            rows = soup.find_all("div", class_="divTableRow")

            for row in rows:

                cells = row.find_all("div", class_="divTableCell")

                if len(cells) != 7:
                    continue

                code = cells[0].get_text(strip=True)

                # Skip Header
                if code == "Subject Code":
                    continue

                # Skip Footer
                if code.startswith("P ->"):
                    continue

                subject = {

                    "code": code,

                    "name": cells[1].get_text(strip=True),

                    "internal": cells[2].get_text(strip=True),

                    "external": cells[3].get_text(strip=True),

                    "total": cells[4].get_text(strip=True),

                    "result": cells[5].get_text(strip=True),

                    "date": cells[6].get_text(strip=True)

                }

                subjects.append(subject)

            student["subjects"] = subjects
            if len(subjects) == 0:

                print(f"⚠️ No subjects found for {usn}. Retrying...")

                time.sleep(2)

                continue

            print("Subjects Found :", len(subjects))

            print(f"✅ Result fetched successfully for {usn}")

            return student

        except Exception as e:

            print(f"⚠️ Error fetching {usn}: {e}")
            print(f"🔄 Retrying {usn}...")

            time.sleep(2)

            continue
# -----------------------
# Save Results to Excel
# -----------------------
def save_results_to_excel(results):

    if not results:
        print("❌ No Results Found")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "VTU Results"

    header = ["USN", "Name"]

    first_student = results[0]

    for subject in first_student["subjects"]:
        header.append(subject["code"])

    header.append("Result")

    ws.append(header)

    red_fill = PatternFill(
        start_color="FF4D4D",
        end_color="FF4D4D",
        fill_type="solid"
    )

    for student in results:

        row = [
            student["USN"],
            student["Name"]
        ]

        final_result = "P"

        for subject in student["subjects"]:
            row.append(subject["total"])

            if subject["result"] == "F":
                final_result = "F"

        row.append(final_result)
        ws.append(row)

        current_row = ws.max_row

        if final_result == "F":
            for cell in ws[current_row]:
                cell.fill = red_fill

    os.makedirs("ExcelFiles", exist_ok=True)
    wb.save("ExcelFiles/results.xlsx")

    print("✅ Excel Saved")


# -----------------------
# Range Scraper Function
# -----------------------
def scrape_range():

    results = []

    for u in range(low, high):

        usn = f"{college}{year}{branch}{str(u).zfill(3)}"

        print(f"\nFetching result for {usn}")

        res = fetch_result(usn)

        if res:
            results.append(res)

    print("Total Results:", len(results))

    save_results_to_excel(results)

    driver.quit()

    print("🎉 SCRAPING COMPLETED")

# -----------------------
# Run Main Range Scraper
# -----------------------
if __name__ == "__main__":

    college, year, branch, low, high, semc, url = get_inputs()

    # --- Set subject code and loop count based on diploma and semester ---
    dip = "Y" if low >= 400 else "N"
    subcode = 52
    iloop = 8

    if semc in ["3", "4"]:
        iloop = 9
        subcode = 58

        if dip == "Y":
            iloop = 10
            subcode = 64

    scrape_range()